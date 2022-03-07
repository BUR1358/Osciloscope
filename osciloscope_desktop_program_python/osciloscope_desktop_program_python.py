import sys
import numpy as np
import pyqtgraph as pg
import serial
from PySide2 import QtWidgets, QtGui
from PySide2.QtCore import *
from PySide2.QtGui import *
from PySide2 import QtCore
from PySide2.QtWidgets import QPushButton, QLineEdit, QLabel, QMessageBox, QComboBox, QFileDialog, QSlider
import openpyxl

dialog_name = 'Сохраниние'
Folder_init = 'C:\Program Files'
level_synch_val = 0
speed_default = 115200
speed = None

delta_default = 15
delta = 15
delta_range = np.arange(0.0, 50.1, 0.01)

read_points = 2048
num_points = 1023

COM_defualt = 'COM4'
COM_port = None

try:
    ser = serial.Serial(COM_port, baudrate=speed, timeout=0.005)
except:
    ser = None
x = []
values_v1 = []
values_v2 = []

values_v1_plot = []
values_v2_plot = []

count_value = 0

y_start = 0
y_stop = 3
y_range_lim = np.arange(0.0, 3.1, 0.01)

x_start = 0
x_stop = 1022
x_range_lim = np.arange(0.0, 1023.1, 0.01)

clicker_value = 0

time_update_default = 0.001
time_update = 0.001

time_update_range = np.arange(0.0, 1000.1, 0.001)


def show_error_message_COM(self, com_val):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText(f'COM-{com_val} port не активен.')
    msg.setIcon(QMessageBox.Warning)
    msg.setInformativeText(f"Выбранный COM-{com_val} port не активен. Выберите другой COM port.")
    x = msg.exec_()


def show_error_message_range_y(self):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText('Вы ввели некорректные пределы по оси-Y.')
    msg.setIcon(QMessageBox.Warning)
    msg.setInformativeText("Пожалуйста, укажите пределы в диапазоне от 0 до 3 В.")
    x = msg.exec_()


def show_error_message_range_x(self):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText('Вы ввели некорректные пределы по оси-X.')
    msg.setIcon(QMessageBox.Warning)
    msg.setInformativeText("Пожалуйста, укажите пределы в диапазоне от 0 до 1023 мс * 10^(-1).")
    x = msg.exec_()


def show_error_message_delta(self):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText('Вы ввели некорректную Delta.')
    msg.setIcon(QMessageBox.Warning)
    msg.setInformativeText("Пожалуйста, укажите значение в диапазоне от 1 до 50.")
    x = msg.exec_()


def show_error_message_time_update(self):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText('Вы ввели некорректное время обновления графика.')
    msg.setIcon(QMessageBox.Warning)
    msg.setInformativeText("Пожалуйста, укажите значение от 0 до 1000 мс.")
    x = msg.exec_()


def show_error_save(self):
    msg = QMessageBox()
    appIcon_error = QIcon("C:\Prog_project\Python\Osciloscope\icon1.ico")
    msg.setWindowIcon(appIcon_error)
    msg.setWindowTitle('Ошибка')
    msg.setText('Сохранение не удалось.')
    msg.setIcon(QMessageBox.Warning)
    x = msg.exec_()


def setIcon(self):  # иконка
    appIcon = QIcon("C:\Prog_project\Python\Osciloscope\icon.ico")
    self.setWindowIcon(appIcon)


class MyWidget(pg.GraphicsWindow, QtWidgets.QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.graph = pg.GraphicsWindow()
        self.setWindowTitle("Oscilloscope")
        setIcon(self)
        self.mainLayout = QtWidgets.QVBoxLayout()
        self.setLayout(self.mainLayout)
        self.grid = QtWidgets.QGridLayout()

        self.button_start = QPushButton('Старт')  # Кнопка построить
        self.button_start.setMinimumSize(300, 40)
        self.button_start.setMaximumSize(300, 40)
        self.button_start.setFocus()
        self.button_start.setStyleSheet("""
                                QPushButton{
                                    font-weight: bold;
                                    border-radius: 7px;
                                    color: #ffffff;
                                    background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #29e100, stop:0.90 #009c3f);
                                    font: 12pt 'Arial'
                                }
                                """)
        # --------------------------------------------------------------------------------------------------------------
        self.button_stop = QPushButton('Стоп')
        self.button_stop.setMinimumSize(300, 40)
        self.button_stop.setMaximumSize(300, 40)
        self.button_stop.setStyleSheet("""
                                        QPushButton{
                                            font-weight: bold;
                                            border-radius: 7px;
                                            color: #ffffff;
                                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #da066d, stop:0.90 #f65f1a);
                                            font: 12pt 'Arial'
                                        }
                                        """)
        # --------------------------------------------------------------------------------------------------------------
        self.button_clear = QPushButton('Очистить')  # Кнопка очистки
        self.button_clear.setMinimumSize(300, 40)
        self.button_clear.setMaximumSize(300, 40)
        self.button_clear.setStyleSheet("""
                                        QPushButton{
                                            font-weight: bold;
                                            border-radius: 7px;
                                            color: #ffffff;
                                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #69040c, stop:0.90 #c60115);
                                            font: 12pt 'Arial'
                                        }
                                        """)
        # --------------------------------------------------------------------------------------------------------------
        self.Save_v1 = QPushButton('Save U1')  # Кнопка очистки
        self.Save_v1.setMinimumSize(147, 30)
        self.Save_v1.setMaximumSize(147, 30)
        self.Save_v1.setStyleSheet("""
                                        QPushButton{
                                            font-weight: bold;
                                            border-radius: 7px;
                                            color: #ffffff;
                                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #1173f3, stop:0.90 #073fc3);
                                            font: 12pt 'Arial'
                                        }
                                        """)
        # --------------------------------------------------------------------------------------------------------------
        self.Save_v2 = QPushButton('Save U2')  # Кнопка очистки
        self.Save_v2.setMinimumSize(147, 30)
        self.Save_v2.setMaximumSize(147, 30)
        self.Save_v2.setStyleSheet("""
                                        QPushButton{
                                            font-weight: bold;
                                            border-radius: 7px;
                                            color: #ffffff;
                                            background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #1173f3, stop:0.90 #073fc3);
                                            font: 12pt 'Arial'
                                        }
                                        """)
        # --------------------------------------------------------------------------------------------------------------
        self.y_start_line = QLineEdit()
        self.y_start_line.setAlignment(Qt.AlignCenter)
        self.y_start_line.setMinimumSize(147, 30)
        self.y_start_line.setMaximumSize(147, 30)
        self.y_start_line.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.y_stop_line = QLineEdit()
        self.y_stop_line.setAlignment(Qt.AlignCenter)
        self.y_stop_line.setMinimumSize(147, 30)
        self.y_stop_line.setMaximumSize(147, 30)
        self.y_stop_line.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.x_start_line = QLineEdit()
        self.x_start_line.setAlignment(Qt.AlignCenter)
        self.x_start_line.setMinimumSize(147, 30)
        self.x_start_line.setMaximumSize(147, 30)
        self.x_start_line.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.x_stop_line = QLineEdit()
        self.x_stop_line.setAlignment(Qt.AlignCenter)
        self.x_stop_line.setMinimumSize(147, 30)
        self.x_stop_line.setMaximumSize(147, 30)
        self.x_stop_line.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.time_update_edit = QLineEdit()
        self.time_update_edit.setAlignment(Qt.AlignCenter)
        self.time_update_edit.setMinimumSize(300, 30)
        self.time_update_edit.setMaximumSize(300, 30)
        self.time_update_edit.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")

        self.delta = QLineEdit()
        self.delta.setAlignment(Qt.AlignCenter)
        self.delta.setMinimumSize(147, 30)
        self.delta.setMaximumSize(147, 30)
        self.delta.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 14pt 'Arial'; border: 1px solid black; border-radius: 7px;")

        # --------------------------------------------------------------------------------------------------------------
        self.speed_label = QLabel('Бит в секунду', self)
        self.speed_label.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.speed_label.setMinimumSize(QSize(147, 30))
        self.speed_label.setMaximumSize(QSize(147, 30))
        self.speed_label.setAlignment(Qt.AlignCenter)

        self.combo_label = QLabel('COM Port', self)
        self.combo_label.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.combo_label.setMinimumSize(QSize(147, 30))
        self.combo_label.setMaximumSize(QSize(147, 30))
        self.combo_label.setAlignment(Qt.AlignCenter)

        self.time_update_label = QLabel('Время обновления графика [мс]', self)
        self.time_update_label.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.time_update_label.setMinimumSize(QSize(300, 30))
        self.time_update_label.setMaximumSize(QSize(300, 30))
        self.time_update_label.setAlignment(Qt.AlignCenter)

        self.value_y_points = QLabel('Пределы по Y-оси [min, max]', self)
        self.value_y_points.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.value_y_points.setMinimumSize(QSize(300, 30))
        self.value_y_points.setMaximumSize(QSize(300, 30))
        self.value_y_points.setAlignment(Qt.AlignCenter)

        self.value_x_points = QLabel('Пределы по X-оси [min, max]', self)  # Кол-во точек подпись
        self.value_x_points.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.value_x_points.setMinimumSize(QSize(300, 30))
        self.value_x_points.setMaximumSize(QSize(300, 30))
        self.value_x_points.setAlignment(Qt.AlignCenter)

        self.slider_laber = QLabel('0 [В]', self)
        self.slider_laber.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.slider_laber.setMinimumSize(QSize(147, 30))
        self.slider_laber.setMaximumSize(QSize(147, 30))
        self.slider_laber.setAlignment((Qt.AlignCenter | Qt.AlignVCenter))

        self.synch = QLabel('СИНХРОНИЗАЦИЯ', self)
        self.synch.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.synch.setMinimumSize(QSize(300, 30))
        self.synch.setMaximumSize(QSize(300, 30))
        self.synch.setAlignment(Qt.AlignCenter)

        self.delta_label = QLabel('Delta', self)
        self.delta_label.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.delta_label.setMinimumSize(QSize(147, 30))
        self.delta_label.setMaximumSize(QSize(147, 30))
        self.delta_label.setAlignment(Qt.AlignCenter)
        # --------------------------------------------------------------------------------------------------------------
        self.slider = QSlider(QtCore.Qt.Horizontal, parent=parent)
        self.slider.setRange(0, 1000)
        self.slider.setPageStep(5)
        self.slider.valueChanged.connect(self.Update_slider)
        self.slider.setMinimumSize(QSize(150, 30))
        self.slider.setMaximumSize(QSize(150, 30))

        # --------------------------------------------------------------------------------------------------------------
        self.combo = QComboBox()
        self.combo.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.combo.setMinimumSize(QSize(147, 30))
        self.combo.setMaximumSize(QSize(147, 30))

        self.combo.addItems(["COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9"])
        # --------------------------------------------------------------------------------------------------------------
        self.speed = QComboBox()
        self.speed.setStyleSheet(
            "background-color: #2c2d2f; color: white; font: 12pt 'Arial'; border: 1px solid black; border-radius: 7px;")
        self.speed.setMinimumSize(QSize(147, 30))
        self.speed.setMaximumSize(QSize(147, 30))

        self.speed.addItems(
            ["4800", "9600", "14400", "19200", "28800", "38400", "56000", "57600", "115200", "128000", "256000"])
        # --------------------------------------------------------------------------------------------------------------
        self.mainLayout.addLayout(self.grid)
        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.onNewData)
        # --------------------------------------------------------------------------------------------------------------
        self.a = self.graph.addPlot(title="U1(t), U2(t)")
        self.graph.setMinimumSize(900, 750)

        self.a.setLabels(
            bottom='Время [мс * 10^(-1)]',
            left='Напряжение [В]')
        self.a.addLegend()

        self.curve1 = self.a.plot()
        self.curve2 = self.a.plot()
        self.a.setYRange(y_start, y_stop)
        self.a.setXRange(x_start, x_stop)
        self.a.showGrid(x=True, y=True)
        self.grid.addWidget(self.graph, 0, 0, 0, 1)
        self.grid.addWidget(self.button_start, 0, 1, 1, 2)
        self.grid.addWidget(self.button_stop, 1, 1, 1, 2)

        self.grid.addWidget(self.combo_label, 2, 1, 1, 1)
        self.grid.addWidget(self.speed_label, 2, 2, 1, 1)

        self.grid.addWidget(self.combo, 3, 1, 1, 1)
        self.grid.addWidget(self.speed, 3, 2, 1, 1)
        self.grid.addWidget(self.value_y_points, 4, 1, 2, 2)
        self.grid.addWidget(self.y_start_line, 5, 1, 1, 1)
        self.grid.addWidget(self.y_stop_line, 5, 2, 1, 1)

        self.grid.addWidget(self.value_x_points, 6, 1, 2, 2)
        self.grid.addWidget(self.x_start_line, 7, 1, 1, 1)
        self.grid.addWidget(self.x_stop_line, 7, 2, 1, 1)

        self.grid.addWidget(self.time_update_label, 8, 1, 2, 2)
        self.grid.addWidget(self.time_update_edit, 9, 1, 1, 2)

        self.grid.addWidget(self.button_clear, 10, 1, 2, 2)
        self.grid.addWidget(self.Save_v1, 11, 1, 2, 2)
        self.grid.addWidget(self.Save_v2, 11, 2, 2, 2)
        self.grid.addWidget(self.synch, 12, 1, 2, 2)
        self.grid.addWidget(self.slider, 13, 1, 1, 2)
        self.grid.addWidget(self.slider_laber, 13, 2, 1, 2)

        self.grid.addWidget(self.delta_label, 14, 1, 1, 2)
        self.grid.addWidget(self.delta, 14, 2, 1, 2)

        # ---Нажатия на кнопки-------------------------------------------------------------------------------------------
        self.button_start.clicked.connect(self.button_start_clicked)  # при нажатии на кнопку перейти к функции
        self.button_stop.clicked.connect(self.button_stop_clicked)
        self.button_clear.clicked.connect(self.button_clear_clicked)
        self.Save_v1.clicked.connect(self.button_Save_v1)
        self.Save_v2.clicked.connect(self.button_Save_v2)

    def button_stop_clicked(self):
        global clicker_value
        clicker_value = 0
        try:
            ser.close()
        except:
            pass
        self.timer.stop()

    def Update_slider(self, level_synch):
        global level_synch_val
        self.slider_laber.setText(str(level_synch * 3 / 1000) + ' [В]')
        level_synch_val = level_synch * 3 / 1000

    def button_start_clicked(self):
        global clicker_value
        global time_update
        global time_update_default
        global speed_default
        global speed
        global COM_port
        global COM_defualt
        global ser, num_points
        if clicker_value == 0:
            clicker_value = 1
            self.a.clear()
            self.timer.start()
            self.a.setMouseEnabled(x=False)
            self.curve1 = self.a.plot(symbol='o', symbolPen=None, symbolSize=5, symbolBrush=('r'), fillLevel=0,
                                      fillBrush=(255, 255, 255, 15), name='input 1 - U1')
            self.curve2 = self.a.plot(symbol='o', symbolPen=None, symbolSize=5, symbolBrush=('y'), fillLevel=0,
                                      fillBrush=(255, 255, 255, 15), name='input 2 - U2')

            y_start = self.y_start_line.text()
            y_stop = self.y_stop_line.text()

            x_start = self.x_start_line.text()
            x_stop = self.x_stop_line.text()

            time_update = self.time_update_edit.text()

            speed = int(self.speed.currentText())
            COM_port = str(self.combo.currentText())
            try:
                ser = serial.Serial(COM_port, baudrate=speed, timeout=0.005)
                try:
                    if float(y_start) in y_range_lim:
                        y_start = float(y_start)
                    else:
                        show_error_message_range_y(self)
                        y_start = 0

                    if float(y_stop) in y_range_lim:
                        y_stop = float(y_stop)
                    else:
                        show_error_message_range_y(self)
                        y_stop = 3

                except ValueError:
                    if y_start or y_stop != '':
                        show_error_message_range_y(self)
                    else:
                        y_start = 0
                        y_stop = 3
                try:
                    x_start = float(x_start)
                    x_stop = float(x_stop)
                except:
                    if x_start or x_stop != '':
                        show_error_message_range_x(self)
                    else:
                        x_start = 0
                        x_stop = 1022

                try:
                    if float(time_update) in time_update_range:
                        time_update = float(time_update)
                    else:
                        show_error_message_time_update(self)
                        time_update = time_update_default

                except ValueError:
                    if time_update != '':
                        show_error_message_time_update(self)
                    else:
                        time_update = time_update_default

                    # ----------------------------------------------------------------------------------------------------------
                try:
                    self.timer.setInterval(time_update)  # in milliseconds
                    self.a.setYRange(y_start, y_stop)
                    self.a.setXRange(x_start, x_stop)
                except:
                    pass


            except serial.SerialException:
                try:
                    ser.close()
                except:
                    pass
                self.timer.stop()
                self.a.clear()
                com_val = COM_port[3:4]
                clicker_value = 0
                show_error_message_COM(self, com_val)
        else:
            return

    def button_clear_clicked(self):
        global clicker_value
        self.timer.stop()
        clicker_value = 0
        self.a.clear()
        try:
            ser.close()
        except:
            pass

    def button_Save_v1(self):
        x = range(len(values_v1_plot))
        filename1 = QFileDialog.getExistingDirectory(self, dialog_name, Folder_init)
        filepath1 = str(filename1) + '/U1.xlsx'
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            sheet = wb.create_sheet("Sheet", -1)
            for i in range(len(x)):
                sheet.cell(row=i + 1, column=1).value = int(x[i])
                sheet.cell(row=i + 1, column=2).value = values_v1_plot[i]
            wb.save(filepath1)
        except:
            show_error_save(self)

    def button_Save_v2(self):
        x = range(len(values_v2_plot))
        filename2 = QFileDialog.getExistingDirectory(self, dialog_name, Folder_init)
        filepath2 = str(filename2) + '/U2.xlsx'
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            sheet = wb.create_sheet("Sheet", -1)
            for i in range(len(x)):
                sheet.cell(row=i + 1, column=1).value = int(x[i])
                sheet.cell(row=i + 1, column=2).value = values_v2_plot[i]
            wb.save(filepath2)
        except:
            show_error_save(self)

    def setData(self, x, values_v1_plot, values_v2_plot):
        try:
            self.curve1.setData(x, values_v1_plot)
            self.curve2.setData(x, values_v2_plot)
        except:
            return

    def onNewData(self):
        global delta, delta_range, delta_default
        global count_value, values_v1, values_v2, values_v1_plot, values_v2_plot, num_points, read_points

        first_element = 0
        sync_val = level_synch_val

        try:
            delta = self.delta.text()
            if float(delta) in delta_range:
                delta = float(delta)
            else:
                show_error_message_delta(self)
                delta = delta_default
                self.timer.stop()
        except ValueError:
            if delta != '':
                show_error_message_delta(self)
                self.timer.stop()
            else:
                delta = delta_default

        try:
            ADC_Data = ser.read(read_points)
            if (int(ADC_Data[0]) == 111) and (int(ADC_Data[1]) == 111):
                for i in range(2, read_points, 2):
                    data = (int(ADC_Data[i]) << 8 | int(ADC_Data[i + 1])) * 3 / 4096
                    values_v1.append(data)
                values_v1_plot = values_v1
                if int(level_synch_val) != 0:
                    for i in range(num_points):
                        if (values_v1_plot[i] >= sync_val) and (values_v1_plot[i] < values_v1_plot[i + delta]):
                            first_element = int(i)
                            break
                    values_v1_plot = values_v1_plot[first_element:]
                values_v1 = []
            if (int(ADC_Data[0]) == 222) and (int(ADC_Data[1]) == 222):
                for i in range(2, read_points, 2):
                    data = (int(ADC_Data[i]) << 8 | int(ADC_Data[i + 1])) * 3 / 4096
                    values_v2.append(data)
                values_v2_plot = values_v2
                if int(level_synch_val) != 0:
                    for z in range(num_points):
                        if (values_v2_plot[z] >= sync_val) and (values_v2_plot[z] < values_v2_plot[z + delta]):
                            first_element = int(z)
                            break
                    values_v2_plot = values_v2_plot[first_element:]
                values_v2 = []

            if int(level_synch_val) != 0:
                if (len(values_v2_plot) and len(values_v1_plot)) != 0:
                    if len(values_v1_plot) >= len(values_v2_plot):
                        values_v1_plot = values_v1_plot[0:len(values_v2_plot)]
                    elif len(values_v2_plot) >= len(values_v1_plot):
                        values_v2_plot = values_v2_plot[0:len(values_v1_plot)]

        except (IndexError, ValueError):
            return
        except (NameError, AttributeError, serial.serialutil.SerialException):
            COM_port = str(self.combo.currentText())
            try:
                ser.close()
            except:
                pass
            self.timer.stop()
            self.a.clear()
            com_val = COM_port[3:4]
            show_error_message_COM(self, com_val)

        x = range(len(values_v2_plot))

        self.setData(x, values_v1_plot, values_v2_plot)


def main():
    app = QtWidgets.QApplication.instance()
    if not app:
        app = QtWidgets.QApplication(sys.argv)
    pg.setConfigOptions(antialias=False)

    win = MyWidget()
    win.show()
    win.resize(1000, 700)
    win.raise_()
    app.exec_()


if __name__ == "__main__":
    main()
